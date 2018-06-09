import openpyxl

wb = openpyxl.Workbook()
wb.get_sheet_names()

sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'].value = 44
sheet['A2'].value = 'Hey...'

import os
os.chdir('')

wb.save('example.xlsx') # to save the excel made in the hard drive

sheet2 = wb.create_sheet()
print(wb.get_sheet_names())

sheet2.title = 'My New Sheet...'
# now the name of the sheet has been changed...
wb.save('example2.xlsx')

# if you want a sheet at a particular index in the Workbook
wb.create_sheet(index = 0, title = 'My other sheet. The Main one now...') # this is the first sheet now...
wb.save('example3.xlsx')
