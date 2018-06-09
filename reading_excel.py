import openpyxl
import os
os.chdir('')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook.get_sheet_by_name('sheet1')
print(type(sheet))

sheet_names = workbook.get_sheet_names()
print(sheet_names) # this is a list containing names of all the sheets

cell = sheet['A1']
print(cell.value)
# columns are alphabetical and rows are numerical... A1, B2, etc...

# for getting using numerical rows as well as columns
sheet.cell(row=1, column=2)

# example
for i in range(1, 8):
	print(sheet.cell(row = i, column=2).value)

